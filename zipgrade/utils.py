"""
ZipGrade CSV Parser Utility

Parses ZipGrade export files (CSV format) and extracts:
- Student information
- Answers and scores
- Question-level data
"""
import csv
import json
import io
from decimal import Decimal
from typing import List, Dict, Any, Tuple, Optional

from schools.utils import normalize_student_id


class ZipGradeParser:
    """Parser for ZipGrade CSV and XLSX export files."""
    
    # Known column mappings based on ZipGrade export format
    STUDENT_ID_COLUMNS = ['ExternalId', 'External ID', 'ZipGrade ID', 'Student ID', 'StudentId']
    FIRST_NAME_COLUMNS = ['FirstName', 'First Name', 'First']
    LAST_NAME_COLUMNS = ['LastName', 'Last Name', 'Last']
    EARNED_COLUMNS = ['EarnedPts', 'Earned Points', 'Earned', 'Points Earned', 'Score']
    MAX_COLUMNS = ['PossiblePts', 'Possible Points', 'Max Points', 'Possible', 'Max']
    PERCENT_COLUMNS = ['Percent', 'Percentage', 'Pct', '%']
    CLASS_COLUMNS = ['Class', 'Section', 'Period', 'Grade']
    
    def __init__(self, file_content: bytes, encoding: str = 'utf-8', filename: str = ''):
        """Initialize parser with file content.
        
        Args:
            file_content: Raw bytes from uploaded file
            encoding: Character encoding (default utf-8)
            filename: Original filename to detect format
        """
        self.file_content = file_content
        self.encoding = encoding
        self.filename = filename.lower()
        self.headers = []
        self.data = []
        self.answer_columns = []
        self.column_map = {}
    
    def _is_xlsx(self) -> bool:
        """Check if file is an XLSX file."""
        return self.filename.endswith('.xlsx') or self.filename.endswith('.xls')
    
    def _parse_xlsx(self) -> Tuple[List[str], List[Dict[str, str]]]:
        """Parse XLSX file and return headers and rows as dicts."""
        from openpyxl import load_workbook
        
        wb = load_workbook(filename=io.BytesIO(self.file_content), read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        
        # First row is headers
        headers = [str(cell) if cell is not None else '' for cell in rows[0]]
        
        # Convert remaining rows to dicts
        data = []
        for row in rows[1:]:
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ''
            data.append(row_dict)
        
        wb.close()
        return headers, data
        
    def parse(self) -> Dict[str, Any]:
        """Parse the ZipGrade file and return structured data.
        
        Returns:
            Dictionary containing:
            - total_questions: Number of questions
            - total_students: Number of students
            - results: List of student results
            - errors: List of parsing errors
        """
        errors = []
        results = []
        
        try:
            # Check if XLSX format
            if self._is_xlsx():
                self.headers, rows_data = self._parse_xlsx()
                if not self.headers:
                    return {
                        'total_questions': 0,
                        'total_students': 0,
                        'results': [],
                        'errors': ['Empty or invalid XLSX file.']
                    }
                
                # Map columns and find answer columns
                self._map_columns()
                self._find_answer_columns()
                
                # Extract answer key from PriKey columns (ZipGrade format)
                answer_key = {}
                if hasattr(self, 'prikey_columns') and self.prikey_columns and rows_data:
                    # PriKey values are same for all rows, so just use first row
                    first_row = rows_data[0]
                    for i, col in enumerate(self.prikey_columns, start=1):
                        answer = str(first_row.get(col, '')).strip().upper()
                        if answer:
                            answer_key[str(i)] = answer
                
                # Parse each row
                for row_num, row in enumerate(rows_data, start=2):
                    try:
                        # Skip Answer Key row if present (legacy format)
                        if self._is_answer_key_row(row):
                            if not answer_key:  # Only extract if not already got from PriKey
                                answer_key = self._extract_answer_key(row)
                            continue
                        
                        result = self._parse_row(row)
                        if result:
                            results.append(result)
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                
                
                # Check for max points to fix possible over-detection of columns
                # If we detected way more columns than max points, trust max points
                derived_total_questions = len(self.answer_columns)  # Initialize with current count
                max_values = [r['max_points'] for r in results if r['max_points'] > 0]
                if max_values:
                    from collections import Counter
                    most_common = Counter(max_values).most_common(1)
                    if most_common:
                        likely_max_points = int(most_common[0][0])
                        
                        # If we found significantly more columns than expected, truncate
                        if len(self.answer_columns) > likely_max_points:
                             # Also update the derived count if needed
                            derived_total_questions = likely_max_points
                            # Truncate answer columns to match likely question count
                            self.answer_columns = self.answer_columns[:likely_max_points]
                            # Also truncate answer key
                            answer_key = {k: v for k, v in answer_key.items() if int(k) <= likely_max_points}

                return {
                    'total_questions': derived_total_questions if derived_total_questions > 0 else len(self.answer_columns),
                    'total_students': len(results),
                    'results': results,
                    'errors': errors,
                    'answer_columns': self.answer_columns,
                    'answer_key': answer_key,
                }
            
            # CSV parsing (original logic)
            # Decode file content
            text_content = self.file_content.decode(self.encoding)
            
            # Try to fix common encoding issues
            if '\ufeff' in text_content:
                text_content = text_content.replace('\ufeff', '')
                
        except UnicodeDecodeError:
            # Try different encodings
            for enc in ['utf-8-sig', 'latin-1', 'cp1251', 'cp1252']:
                try:
                    text_content = self.file_content.decode(enc)
                    self.encoding = enc
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return {
                    'total_questions': 0,
                    'total_students': 0,
                    'results': [],
                    'errors': ['Could not decode file. Please ensure it is a valid CSV or XLSX file.']
                }
        
        # Parse CSV
        reader = csv.DictReader(io.StringIO(text_content))
        self.headers = reader.fieldnames or []
        
        if not self.headers:
            return {
                'total_questions': 0,
                'total_students': 0,
                'results': [],
                'errors': ['Empty or invalid CSV file.']
            }
        
        # Map columns
        self._map_columns()
        
        # Find answer columns (usually named like Q1, Q2, etc. or 1, 2, 3)
        self._find_answer_columns()
        
        # Parse each row
        for row_num, row in enumerate(reader, start=2):
            try:
                result = self._parse_row(row)
                if result:
                    results.append(result)
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        
        # Calculate derived total questions from max points if no columns found
        derived_total_questions = len(self.answer_columns)
        if derived_total_questions == 0 and results:
            # Try to infer from max points (assuming 1 point per question)
            max_values = [r['max_points'] for r in results if r['max_points'] > 0]
            if max_values:
                # Use the most common max score or simply the max
                from collections import Counter
                most_common = Counter(max_values).most_common(1)
                if most_common:
                    derived_total_questions = int(most_common[0][0])
        
        return {
            'total_questions': derived_total_questions,
            'total_students': len(results),
            'results': results,
            'errors': errors,
            'answer_columns': self.answer_columns,
        }
    
    def _map_columns(self):
        """Map column names to standardized names."""
        for header in self.headers:
            header_lower = header.lower().strip()
            
            # Student ID
            for col in self.STUDENT_ID_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['student_id'] = header
                    break
            
            # First name
            for col in self.FIRST_NAME_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['first_name'] = header
                    break
            
            # Last name
            for col in self.LAST_NAME_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['last_name'] = header
                    break
            
            # Earned points
            for col in self.EARNED_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['earned'] = header
                    break
            
            # Max points
            for col in self.MAX_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['max'] = header
                    break
            
            # Percentage
            for col in self.PERCENT_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['percent'] = header
                    break
            
            # Class
            for col in self.CLASS_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['class'] = header
                    break
    
    
    def _find_answer_columns(self):
        """Find columns that contain answer data.
        
        ZipGrade format uses:
        - Stu1, Stu2, ... for student answers
        - PriKey1, PriKey2, ... for correct answers (Answer Key)
        """
        import re
        
        # First, check for ZipGrade format (StuN and PriKeyN columns)
        stu_columns = []
        prikey_columns = []
        
        for header in self.headers:
            if not header:
                continue
            header_str = str(header).strip()
            
            # Match Stu1, Stu2, ... (student answers)
            if re.match(r'^Stu\d+$', header_str):
                stu_columns.append(header_str)
            
            # Match PriKey1, PriKey2, ... (correct answers / answer key)
            elif re.match(r'^PriKey\d+$', header_str):
                prikey_columns.append(header_str)
        
        # Sort by number
        def get_num(col):
            match = re.search(r'(\d+)$', col)
            return int(match.group(1)) if match else 999999
        
        stu_columns.sort(key=get_num)
        prikey_columns.sort(key=get_num)
        
        # If we found ZipGrade format, use it
        if stu_columns:
            self.answer_columns = stu_columns
            self.prikey_columns = prikey_columns
            return
        
        # Fallback to old detection for other formats
        self.prikey_columns = []
        mapped_headers = set(self.column_map.values())
        candidates = []
        
        for header in self.headers:
            if header in mapped_headers:
                continue
            header_stripped = str(header).strip()
            upper_header = header_stripped.upper()
            
            if upper_header in ['DATE', 'TIME', 'SCHOOL', 'CLASS', 'SECTION', 'TEACHER', 'SUBJECT', 'EXAM']:
                continue
            
            # Q1, Q2, etc.
            if re.match(r'^Q\s*[-_]?\s*\d+', upper_header):
                candidates.append(header)
                continue
            
            # Just a digit
            if header_stripped.isdigit():
                candidates.append(header)
                continue
            
            # Ends with digit
            if re.search(r'(\d+)$', header_stripped):
                candidates.append(header)
                continue
        
        candidates.sort(key=get_num)
        self.answer_columns = candidates
    
    def _is_answer_key_row(self, row: Dict[str, str]) -> bool:
        """Check if this row contains the Answer Key.
        
        ZipGrade exports have a row where the name/ID contains 'KEY' or 'ANSWER'.
        This row contains the correct answers for each question.
        """
        # Check student ID field
        if 'student_id' in self.column_map:
            student_id = str(row.get(self.column_map['student_id'], '')).strip().upper()
            if 'KEY' in student_id or 'ANSWER' in student_id or 'ОТВЕТ' in student_id:
                return True
        
        # Check name fields
        if 'first_name' in self.column_map:
            first_name = str(row.get(self.column_map['first_name'], '')).strip().upper()
            if 'KEY' in first_name or 'ANSWER' in first_name or 'ОТВЕТ' in first_name:
                return True
        
        if 'last_name' in self.column_map:
            last_name = str(row.get(self.column_map['last_name'], '')).strip().upper()
            if 'KEY' in last_name or 'ANSWER' in last_name or 'КЛЮЧ' in last_name:
                return True
        
        return False
    
    def _extract_answer_key(self, row: Dict[str, str]) -> Dict[str, str]:
        """Extract answer key from the KEY row.
        
        Returns:
            Dictionary mapping question number to correct answer: {"1": "A", "2": "B", ...}
        """
        answer_key = {}
        for i, col in enumerate(self.answer_columns, start=1):
            answer = str(row.get(col, '')).strip().upper()
            if answer:  # Only include non-empty answers
                answer_key[str(i)] = answer
        return answer_key
    
    def _parse_row(self, row: Dict[str, str]) -> Optional[Dict[str, Any]]:
        """Parse a single row of data.
        
        Args:
            row: Dictionary of column name -> value
            
        Returns:
            Parsed result dictionary or None if row should be skipped
        """
        # Get student ID
        student_id = ''
        if 'student_id' in self.column_map:
            student_id = str(row.get(self.column_map['student_id'], '')).strip()
        
        # Skip empty rows
        if not student_id and not any(row.values()):
            return None
        
        # Use a placeholder for missing student IDs
        if not student_id:
            student_id = 'NO_ID'
        
        # Get names
        first_name = ''
        if 'first_name' in self.column_map:
            first_name = str(row.get(self.column_map['first_name'], '')).strip()
        
        last_name = ''
        if 'last_name' in self.column_map:
            last_name = str(row.get(self.column_map['last_name'], '')).strip()
        
        # Get scores
        earned = Decimal('0')
        if 'earned' in self.column_map:
            try:
                earned_str = str(row.get(self.column_map['earned'], '0')).strip()
                earned = Decimal(earned_str.replace(',', '.')) if earned_str else Decimal('0')
            except:
                pass
        
        max_points = Decimal('0')
        if 'max' in self.column_map:
            try:
                max_str = str(row.get(self.column_map['max'], '0')).strip()
                max_points = Decimal(max_str.replace(',', '.')) if max_str else Decimal('0')
            except:
                pass
        
        # Calculate percentage
        percentage = Decimal('0')
        if 'percent' in self.column_map:
            try:
                pct_str = str(row.get(self.column_map['percent'], '0')).strip()
                pct_str = pct_str.replace('%', '').replace(',', '.')
                percentage = Decimal(pct_str) if pct_str else Decimal('0')
            except:
                pass
        elif max_points > 0:
            percentage = (earned / max_points) * 100
        
        # Get class/section
        class_name = ''
        if 'class' in self.column_map:
            class_name = str(row.get(self.column_map['class'], '')).strip()
        
        # Get answers
        answers = {}
        for i, col in enumerate(self.answer_columns, start=1):
            answer = str(row.get(col, '')).strip().upper()
            answers[str(i)] = answer
        
        return {
            'student_id': student_id,
            'student_id_normalized': normalize_student_id(student_id),
            'first_name': first_name,
            'last_name': last_name,
            'earned': float(earned),
            'max_points': float(max_points),
            'percentage': float(round(percentage, 2)),
            'class_name': class_name,
            'answers': answers,
        }


def calculate_subject_scores(
    answers: Dict[str, str],
    answer_key: Dict[str, str],
    subject_splits: List[Dict],
) -> List[Dict]:
    """Calculate per-subject scores based on question ranges.
    
    Args:
        answers: Student's answers {question_num: answer}
        answer_key: Correct answers {question_num: correct_answer}
        subject_splits: List of {'subject_id': int, 'start': int, 'end': int, 'points': float}
        
    Returns:
        List of subject scores
    """
    results = []
    
    for split in subject_splits:
        start_q = split['start']
        end_q = split['end']
        points = split.get('points', 1.0)
        
        correct = 0
        total = 0
        question_results = {}
        
        for q_num in range(start_q, end_q + 1):
            q_str = str(q_num)
            student_answer = answers.get(q_str, '')
            correct_answer = answer_key.get(q_str, '')
            
            total += 1
            is_correct = student_answer == correct_answer if correct_answer else False
            question_results[q_str] = {
                'answer': student_answer,
                'correct': correct_answer,
                'is_correct': is_correct
            }
            
            if is_correct:
                correct += 1
        
        earned = correct * points
        max_pts = total * points
        pct = (earned / max_pts * 100) if max_pts > 0 else 0
        
        results.append({
            'subject_id': split['subject_id'],
            'subject_split_id': split.get('split_id'),
            'earned': earned,
            'max_points': max_pts,
            'percentage': round(pct, 2),
            'correct_count': correct,
            'total_count': total,
            'question_results': question_results,
        })
    
    return results
